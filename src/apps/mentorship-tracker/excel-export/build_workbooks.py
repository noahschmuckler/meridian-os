#!/usr/bin/env python3
"""build_workbooks.py — generates the four regional Onboarding Tracker Excel
workbooks from curriculum-data.json (produced by extract-data.mjs).

Usage:
    node extract-data.mjs
    python3 build_workbooks.py [output_dir]

One workbook per Care Delivery Organization region. Hudson Valley East ships
with the demo cohort (checks + scores seeded) so dashboards render alive; the
other regions ship with an example roster row and empty tracking surfaces.
"""

import json
import sys
import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

HERE = Path(__file__).parent
DATA = json.loads((HERE / "curriculum-data.json").read_text())
OUT_DIR = Path(sys.argv[1]) if len(sys.argv) > 1 else HERE / "out"
OUT_DIR.mkdir(parents=True, exist_ok=True)

TODAY = datetime.date.today()

# ─── Region registry (Care Delivery Organizations) ──────────────────────────
REGIONS = [
    dict(id="hve", name="Hudson Valley East", legacy="CareMount Medical",
         director="Dr. Scott Freiberg", accent="028090", seed=True),
    dict(id="hvw", name="Hudson Valley West", legacy="Crystal Run Healthcare",
         director="Dr. Noah Schmuckler", accent="7C3AED", seed=False),
    dict(id="nj", name="New Jersey", legacy="Riverside Medical Group",
         director="Dr. Raj Brahmbhatt", accent="C2410C", seed=False),
    dict(id="li", name="Long Island", legacy="ProHealth",
         director="Dr. Pauline Mizrachi", accent="059669", seed=False),
]
PROGRAM_ADMINS = ("Program Administrators: Dr. Scott Freiberg (Hudson Valley East) "
                  "and Dr. Noah Schmuckler (Hudson Valley West)")

# ─── Palette / shared styles ─────────────────────────────────────────────────
NAVY = "0F1B2D"
MONDRIAN = ["C73E33", "2B5DA8", "D9A521", "EDE8DF"]  # red, blue, gold, off-white
FONT = "Arial"
N_SLOTS = 12          # provider slots per region
ROSTER_FIRST_ROW = 5  # first roster data row

def F(sz=10, bold=False, color="1F2937", italic=False):
    return Font(name=FONT, size=sz, bold=bold, color=color, italic=italic)

def fill(hex6):
    return PatternFill("solid", start_color=hex6)

THIN = Side(style="thin", color="D1D5DB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
UNLOCKED = Protection(locked=False)

INPUT_FILL = fill("FFFBD6")   # pale yellow = fill me in
GREEN_FONT = Font(name=FONT, size=10, bold=True, color="15803D")

# ─── Curriculum data ─────────────────────────────────────────────────────────
MC = DATA["master_checklist"]
MD_ITEMS = MC["items"]
MD_PHASE_LABEL = {p["id"]: p["label"] for p in MC["phases"]}
PHYS = DATA["mentor_physician_track"]
APC = DATA["mentor_apc_track"]
ALL_MENTOR_PHASES = PHYS + APC                     # w0 … m24, in order
OM = DATA["office_manager_reviews"]
QP = DATA["provider_questionnaires"]
PROVS = DATA["seed_providers"]
USERS = {u["id"]: u["name"] for u in DATA["seed_users"]}
PRODM = DATA["productivity_metrics"]
SCORES = DATA["seed_scores"]
OM_SCORES = DATA["seed_om_scores"]
CII_SCORES = DATA["seed_cii_scores"]

# Nominal start day for each mentor phase — drives Expected Phase / Status.
PHASE_DAYS = {"w0": 0, "w1": 7, "w2": 14, "w3": 21, "m1": 30, "w5": 35,
              "w6": 42, "w7": 49, "m2": 60, "m3": 90, "m4": 120, "m5": 150,
              "m6": 180, "m7": 210, "m8": 240, "m9": 270, "m12": 365,
              "m15": 455, "m18": 545, "m21": 635, "m24": 730}
PHASE_IDS = [p["id"] for p in ALL_MENTOR_PHASES]
PHASE_LABELS = [p["label"] for p in ALL_MENTOR_PHASES]
N_PHASES = len(PHASE_IDS)
PHASE_LIST_RANGE = f"Lists!$B$2:$B${1 + N_PHASES}"
DAY_LIST_RANGE = f"Lists!$C$2:$C${1 + N_PHASES}"

METRIC_COLS = [  # (header, unit, goal, better, description)
    ("Patients per Day", "patients", 18, "Higher is better",
     "Average completed patient visits per clinic day for the week."),
    ("RVUs per Week", "wRVU", 185, "Higher is better",
     "Total work RVUs credited for the week."),
    ("In-Basket Minutes per Day", "minutes", 22, "Lower is better",
     "Average minutes per day spent in the Epic In-Basket."),
    ("Minutes per Note", "minutes", 14, "Lower is better",
     "Average documentation time per progress note."),
    ("Refill Turnaround Minutes", "minutes", 6.5, "Lower is better",
     "Average time from refill request arrival to completed action."),
]

SHEETS = dict(
    home="Home", roster="Roster", dash="Dashboard",
    md="Medical Director Curriculum",
    phys="Mentor - Physician Track", apc="Mentor - APC Track",
    om="Office Manager Reviews", q="Provider Questionnaires",
    metrics="Provider Metrics", dd="Data Dictionary", lists="Lists",
)

def q(sheet):  # quoted sheet name for formulas
    return f"'{sheet}'"

def prov_col(slot, first_col):
    """Column letter for provider slot 1..N_SLOTS starting at first_col letter."""
    return get_column_letter(ord(first_col) - ord("A") + slot)

def home_link(ws):
    c = ws["A1"]
    c.value = "⌂ Home"
    c.font = Font(name=FONT, size=10, bold=True, color="2B5DA8", underline="single")
    c.hyperlink = Hyperlink(ref="A1", location=f"{q(SHEETS['home'])}!A1", tooltip="Back to Home")

def title_block(ws, region, subtitle):
    ws["A2"] = ws.title
    ws["A2"].font = F(14, bold=True, color=NAVY)
    ws["B2"] = f"{region['name']} · {subtitle}"
    ws["B2"].font = F(10, color="6B7280", italic=True)

def header_formula(slot):
    r = ROSTER_FIRST_ROW + slot - 1
    return f'=IF({q(SHEETS["roster"])}!$B${r}="","",{q(SHEETS["roster"])}!$B${r})'

def style_header_cell(cell, accent):
    cell.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
    cell.fill = fill(accent)
    cell.alignment = CENTER_WRAP
    cell.border = BOX

def protect(ws):
    ws.protection.sheet = True
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False

CHECK_DV = lambda: DataValidation(type="list", formula1='"✔"', allow_blank=True,
                                  showErrorMessage=True,
                                  error="Type ✔ (or pick from the list) to mark complete; delete to clear.")
SCORE_DV = lambda: DataValidation(type="whole", operator="between", formula1="1",
                                  formula2="10", allow_blank=True, showErrorMessage=True,
                                  error="Scores are whole numbers from 1 to 10.")

# ═════════════════════════════════════════════════════════════════════════════
def build_lists(wb):
    ws = wb.create_sheet(SHEETS["lists"])
    ws["A1"], ws["B1"], ws["C1"] = "Phase ID", "Phase Label", "Nominal Start Day"
    for i, pid in enumerate(PHASE_IDS):
        r = 2 + i
        ws[f"A{r}"] = pid
        ws[f"B{r}"] = PHASE_LABELS[i]
        ws[f"C{r}"] = PHASE_DAYS[pid]
    ws["E1"] = "Credentials"
    for i, c in enumerate(["MD", "DO", "NP", "PA"]):
        ws[f"E{2 + i}"] = c
    ws.sheet_state = "hidden"
    return ws

# ═════════════════════════════════════════════════════════════════════════════
def build_roster(wb, region):
    ws = wb.create_sheet(SHEETS["roster"])
    home_link(ws)
    title_block(ws, region, "provider roster — one row per onboarding provider")
    ws["A3"] = ("Fill the yellow cells (ID, Name, Credential, Mentor, Start Date, Current Phase, Notes). "
                "Days / Expected Phase / Status compute automatically.")
    ws["A3"].font = F(9, color="6B7280", italic=True)

    headers = ["Provider ID", "Provider Name", "Credential", "Mentor", "Start Date",
               "Days in Practice", "Current Phase", "Expected Phase", "Status", "Notes"]
    widths = [11, 24, 10, 18, 12, 10, 13, 13, 11, 40]
    hr = ROSTER_FIRST_ROW - 1
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=hr, column=i, value=h)
        style_header_cell(cell, region["accent"])
        ws.column_dimensions[get_column_letter(i)].width = w

    cred_dv = DataValidation(type="list", formula1="Lists!$E$2:$E$5", allow_blank=True)
    phase_dv = DataValidation(type="list", formula1=PHASE_LIST_RANGE, allow_blank=True)
    ws.add_data_validation(cred_dv)
    ws.add_data_validation(phase_dv)

    for slot in range(N_SLOTS):
        r = ROSTER_FIRST_ROW + slot
        ws[f"F{r}"] = f'=IF($E{r}="","",TODAY()-$E{r})'
        ws[f"H{r}"] = (f'=IF(OR($E{r}="",$G{r}=""),"",'
                       f'INDEX({PHASE_LIST_RANGE},MATCH($F{r},{DAY_LIST_RANGE},1)))')
        ws[f"I{r}"] = (f'=IF(OR($E{r}="",$G{r}=""),"",'
                       f'IF(MATCH($G{r},{PHASE_LIST_RANGE},0)>=MATCH($F{r},{DAY_LIST_RANGE},1),"On Track",'
                       f'IF(MATCH($F{r},{DAY_LIST_RANGE},1)-MATCH($G{r},{PHASE_LIST_RANGE},0)=1,"Due","Overdue")))')
        ws[f"E{r}"].number_format = "m/d/yyyy"
        for col in "ABCDEGJ":
            c = ws[f"{col}{r}"]
            c.protection = UNLOCKED
            c.fill = INPUT_FILL
        for col in "ABCDEFGHIJ":
            ws[f"{col}{r}"].border = BOX
        cred_dv.add(ws[f"C{r}"])
        phase_dv.add(ws[f"G{r}"])

    last = ROSTER_FIRST_ROW + N_SLOTS - 1
    status_rng = f"I{ROSTER_FIRST_ROW}:I{last}"
    ws.conditional_formatting.add(status_rng, CellIsRule(
        operator="equal", formula=['"Overdue"'], fill=fill("FECACA"),
        font=Font(name=FONT, bold=True, color="B91C1C")))
    ws.conditional_formatting.add(status_rng, CellIsRule(
        operator="equal", formula=['"Due"'], fill=fill("FEF3C7"),
        font=Font(name=FONT, bold=True, color="B45309")))
    ws.conditional_formatting.add(status_rng, CellIsRule(
        operator="equal", formula=['"On Track"'], fill=fill("DCFCE7"),
        font=Font(name=FONT, color="15803D")))

    if region["seed"]:
        label_of = {p["id"]: p["label"] for p in ALL_MENTOR_PHASES}
        for slot, p in enumerate(PROVS):
            r = ROSTER_FIRST_ROW + slot
            ws[f"A{r}"] = p["id"]
            ws[f"B{r}"] = p["name"]
            ws[f"C{r}"] = p["role"]
            ws[f"D{r}"] = USERS.get(p["mentor"], "")
            ws[f"E{r}"] = TODAY - datetime.timedelta(days=p["days"])
            ws[f"E{r}"].number_format = "m/d/yyyy"
            ws[f"G{r}"] = label_of[p["phase"]]
        ws[f"J{ROSTER_FIRST_ROW}"] = "Demo cohort — replace with your live providers."
    else:
        r = ROSTER_FIRST_ROW
        ws[f"A{r}"], ws[f"B{r}"], ws[f"C{r}"] = "EX-01", "Dr. A. Example", "MD"
        ws[f"D{r}"] = "(assigned mentor)"
        ws[f"E{r}"] = TODAY - datetime.timedelta(days=9)
        ws[f"E{r}"].number_format = "m/d/yyyy"
        ws[f"G{r}"] = "Week 1"
        ws[f"J{r}"] = "EXAMPLE ROW — replace with your first real provider."
        ws[f"J{r}"].font = F(9, color="B45309", italic=True)

    ws.freeze_panes = f"A{ROSTER_FIRST_ROW}"
    protect(ws)
    return ws

# ═════════════════════════════════════════════════════════════════════════════
def checked_mentor_phases(prov):
    """Seed rule from the app: every mentor phase up to and including the
    provider's current phase is complete."""
    cur = PHASE_IDS.index(prov["phase"])
    return set(PHASE_IDS[: cur + 1])

def nonblank_count(rng):
    """Count filled cells in a range of *typed* values (never formulas — COUNTA
    counts a formula returning "" as filled). Used for the roster name column
    and the metrics provider-ID column, both of which are hand/feed entered."""
    return f"COUNTA({rng})"

def build_curriculum_sheet(wb, region, key, title_note, blocks, first_prov_col,
                           seeded_check, credential_gate=None):
    """Generic checklist grid: blocks = [(header_text, [row_label_cols...])].
    Rows: 1 home, 2 title, 3 completion stats, 5 header, 6+ items."""
    ws = wb.create_sheet(SHEETS[key])
    home_link(ws)
    title_block(ws, region, title_note)

    item_col = chr(ord(first_prov_col) - 1)          # item text sits just left of providers
    n_items = sum(len(items) for _, items in blocks)
    first_item_row = ROSTER_FIRST_ROW + 1            # row 6
    # count rows: each block has 1 header row + item rows
    last_row = first_item_row + len(blocks) + n_items - 1

    ws[f"{item_col}3"] = "Completion"
    ws[f"{item_col}3"].font = F(9, bold=True, color="6B7280")
    ws[f"{item_col}3"].alignment = Alignment(horizontal="right")

    dv = CHECK_DV()
    ws.add_data_validation(dv)

    for slot in range(1, N_SLOTS + 1):
        col = prov_col(slot, first_prov_col)
        hcell = ws[f"{col}{ROSTER_FIRST_ROW}"]
        hcell.value = header_formula(slot)
        style_header_cell(hcell, region["accent"])
        ws.column_dimensions[col].width = 13
        rng = f"{col}{first_item_row}:{col}{last_row}"
        # A track that doesn't apply to this provider (e.g. the APC track for a
        # physician) reads blank rather than 0%, so it neither implies unfinished
        # work nor drags the cross-provider averages down.
        blank_when = f'{col}${ROSTER_FIRST_ROW}=""'
        if credential_gate:
            cred = f'{q(SHEETS["roster"])}!$C${ROSTER_FIRST_ROW + slot - 1}'
            tests = ",".join(f'{cred}<>"{c}"' for c in credential_gate)
            blank_when = f"OR({blank_when},AND({tests}))"
        ws[f"{col}3"] = (f'=IF({blank_when},"",'
                         f'COUNTIF({rng},"✔")/{n_items})')
        ws[f"{col}3"].number_format = "0%"
        ws[f"{col}3"].font = F(9, bold=True, color=NAVY)
        ws[f"{col}3"].alignment = CENTER
        dv.add(rng)
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=['"✔"'], fill=fill("DCFCE7"), font=GREEN_FONT))

    r = first_item_row
    for header_text, items in blocks:
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=ord(first_prov_col) - ord("A") + N_SLOTS)
        hc = ws.cell(row=r, column=1, value=header_text)
        hc.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        hc.fill = fill(NAVY)
        hc.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 18
        r += 1
        for row_vals, seeded_slots in items:
            for ci, v in enumerate(row_vals, start=1):
                c = ws.cell(row=r, column=ci, value=v)
                c.font = F(9)
                c.alignment = WRAP
                c.border = BOX
            for slot in range(1, N_SLOTS + 1):
                cc = ws[f"{prov_col(slot, first_prov_col)}{r}"]
                cc.protection = UNLOCKED
                cc.alignment = CENTER
                cc.border = BOX
                if seeded_check and slot in seeded_slots:
                    cc.value = "✔"
                    cc.font = GREEN_FONT
            r += 1

    ws.freeze_panes = f"{first_prov_col}{first_item_row}"
    protect(ws)
    return ws

def build_md_curriculum(wb, region):
    # Group by phase. The master checklist stores items in authoring order, not
    # phase order (Week 1 items appear in six separate runs), so gather all
    # items per phase and emit phases in canonical order — one block per phase,
    # matching how the app builds MD_ITEMS_BY_PHASE.
    by_phase = {}
    for it in MD_ITEMS:
        by_phase.setdefault(it["phase"], []).append(it)
    seeded = set(range(1, len(PROVS) + 1)) if region["seed"] else set()
    blocks = []
    for ph in MC["phases"]:
        items = by_phase.get(ph["id"])
        if not items:
            continue
        rows = []
        for it in items:
            owner = it["owner"] + (f" + {it['partner']}" if it.get("partner") else "")
            rows.append(([it["n"], owner, it.get("section", ""), it["text"]], seeded))
        blocks.append((ph["label"], rows))

    ws = build_curriculum_sheet(
        wb, region, "md",
        "one column per provider — ✔ items as the Medical Director completes them",
        blocks, "E", region["seed"])
    for col, w, h in [("A", 5, "#"), ("B", 12, "Owner"), ("C", 22, "Section"), ("D", 64, "Checklist Item")]:
        ws.column_dimensions[col].width = w
        cell = ws[f"{col}{ROSTER_FIRST_ROW}"]
        cell.value = h
        style_header_cell(cell, region["accent"])
    return ws

def build_mentor_sheet(wb, region, key, phases, note, applies, credential_gate=None):
    seeded_by_prov = {}
    if region["seed"]:
        for slot, p in enumerate(PROVS, start=1):
            if applies(p):
                seeded_by_prov[slot] = checked_mentor_phases(p)
    blocks = []
    for ph in phases:
        items = []
        for i, text in enumerate(ph["items"], start=1):
            slots = {s for s, phs in seeded_by_prov.items() if ph["id"] in phs}
            items.append(([f"{ph['id']}.{i}", text], slots))
        blocks.append((ph["label"], items))
    ws = build_curriculum_sheet(wb, region, key, note, blocks, "C", region["seed"],
                                credential_gate=credential_gate)
    for col, w, h in [("A", 8, "Ref"), ("B", 70, "Check-in Item")]:
        ws.column_dimensions[col].width = w
        cell = ws[f"{col}{ROSTER_FIRST_ROW}"]
        cell.value = h
        style_header_cell(cell, region["accent"])
    return ws

# ═════════════════════════════════════════════════════════════════════════════
def build_question_sheet(wb, region, key, checkpoints, note, seed_lookup,
                         culture_aware):
    """Shared builder for Office Manager Reviews + Provider Questionnaires.
    Hidden helper cols: A = checkpoint id, B = row type (s / t / ci).
    Visible: C label, D question, E.. provider columns."""
    ws = wb.create_sheet(SHEETS[key])
    home_link(ws)
    title_block(ws, region, note)
    first_prov_col = "E"
    first_row = ROSTER_FIRST_ROW + 1

    # Pre-compute layout to know the last row.
    n_rows = sum(1 + len(cp["qs"]) + 1 + (1 if culture_aware and any(qq.get("culture") for qq in cp["qs"]) else 0)
                 for cp in checkpoints)
    last_row = first_row + n_rows - 1

    ws["D3"] = "Average score (all checkpoints)"
    ws["D3"].font = F(9, bold=True, color="6B7280")
    ws["D3"].alignment = Alignment(horizontal="right")
    if culture_aware:
        ws["D4"] = "Culture Integration Index (average of culture items)"
        ws["D4"].font = F(9, bold=True, color="7C3AED")
        ws["D4"].alignment = Alignment(horizontal="right")

    score_dv = SCORE_DV()
    ws.add_data_validation(score_dv)

    for slot in range(1, N_SLOTS + 1):
        col = prov_col(slot, first_prov_col)
        hcell = ws[f"{col}{ROSTER_FIRST_ROW}"]
        hcell.value = header_formula(slot)
        style_header_cell(hcell, region["accent"])
        ws.column_dimensions[col].width = 13
        body = f"{col}{first_row}:{col}{last_row}"
        typ = f"$B{first_row}:$B{last_row}"
        ws[f"{col}3"] = f'=IFERROR(AVERAGEIF({typ},"s",{body}),"")'
        ws[f"{col}3"].number_format = "0.0"
        ws[f"{col}3"].font = F(9, bold=True, color=NAVY)
        ws[f"{col}3"].alignment = CENTER
        if culture_aware:
            ws[f"{col}4"] = f'=IFERROR(AVERAGEIF({typ},"ci",{body}),"")'
            ws[f"{col}4"].number_format = "0.0"
            ws[f"{col}4"].font = F(9, bold=True, color="7C3AED")
            ws[f"{col}4"].alignment = CENTER
        # Color scale over score cells (text/blank cells are ignored).
        ws.conditional_formatting.add(body, ColorScaleRule(
            start_type="num", start_value=1, start_color="F87171",
            mid_type="num", mid_value=6, mid_color="FDE68A",
            end_type="num", end_value=10, end_color="86EFAC"))

    if culture_aware:
        cii_rng = f"E4:{prov_col(N_SLOTS, first_prov_col)}4"
        ws.conditional_formatting.add(cii_rng, CellIsRule(
            operator="lessThan", formula=["6"], fill=fill("FECACA"),
            font=Font(name=FONT, bold=True, color="B91C1C")))

    r = first_row
    for cp in checkpoints:
        ws.merge_cells(start_row=r, start_column=3, end_row=r,
                       end_column=4 + N_SLOTS)
        hc = ws.cell(row=r, column=3, value=cp["label"])
        hc.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        hc.fill = fill(NAVY)
        hc.alignment = Alignment(vertical="center")
        ws[f"A{r}"] = cp["id"]
        ws.row_dimensions[r].height = 18
        r += 1
        block_first = r
        for qq in cp["qs"]:
            is_culture = bool(qq.get("culture"))
            typ = "ci" if (culture_aware and is_culture) else ("s" if qq["ty"] == "s" else "t")
            ws[f"A{r}"] = cp["id"]
            ws[f"B{r}"] = typ
            label = qq.get("label") or ("Culture item" if is_culture else f"Q{qq['qid'].upper()}")
            ws[f"C{r}"] = label
            ws[f"C{r}"].font = F(9, bold=True, color="7C3AED" if is_culture else "374151")
            ws[f"C{r}"].alignment = WRAP
            text = qq["text"]
            if qq["ty"] == "s" and qq.get("anchor_low"):
                text += f"\n(1 = {qq['anchor_low']} · 10 = {qq['anchor_high']})"
            elif qq["ty"] == "s":
                text += "\n(score 1–10)"
            ws[f"D{r}"] = text
            ws[f"D{r}"].font = F(9)
            ws[f"D{r}"].alignment = WRAP
            for slot in range(1, N_SLOTS + 1):
                cc = ws[f"{prov_col(slot, first_prov_col)}{r}"]
                cc.protection = UNLOCKED
                cc.border = BOX
                cc.alignment = CENTER_WRAP if qq["ty"] == "t" else CENTER
                if qq["ty"] == "s":
                    score_dv.add(cc)
                seeded = seed_lookup(slot, cp, qq) if region["seed"] else None
                if seeded is not None:
                    cc.value = seeded
                    if qq["ty"] == "t":
                        cc.font = F(8, italic=True, color="6B7280")
            for col in "CD":
                ws[f"{col}{r}"].border = BOX
            r += 1
        # Checkpoint average row(s)
        scale_cells = [i for i, qq in enumerate(cp["qs"]) if qq["ty"] == "s" and not (culture_aware and qq.get("culture"))]
        ci_cells = [i for i, qq in enumerate(cp["qs"]) if culture_aware and qq.get("culture")]
        ws[f"D{r}"] = "Checkpoint average"
        ws[f"D{r}"].font = F(9, bold=True, italic=True, color="6B7280")
        ws[f"D{r}"].alignment = Alignment(horizontal="right")
        for slot in range(1, N_SLOTS + 1):
            col = prov_col(slot, first_prov_col)
            cells = ",".join(f"{col}{block_first + i}" for i in scale_cells)
            cc = ws[f"{col}{r}"]
            cc.value = f'=IFERROR(AVERAGE({cells}),"")'
            cc.number_format = "0.0"
            cc.font = F(9, bold=True, italic=True, color="6B7280")
            cc.alignment = CENTER
        r += 1
        if ci_cells:
            ws[f"D{r}"] = "Culture Integration Index — this checkpoint"
            ws[f"D{r}"].font = F(9, bold=True, italic=True, color="7C3AED")
            ws[f"D{r}"].alignment = Alignment(horizontal="right")
            for slot in range(1, N_SLOTS + 1):
                col = prov_col(slot, first_prov_col)
                cells = ",".join(f"{col}{block_first + i}" for i in ci_cells)
                cc = ws[f"{col}{r}"]
                cc.value = f'=IFERROR(AVERAGE({cells}),"")'
                cc.number_format = "0.0"
                cc.font = F(9, bold=True, italic=True, color="7C3AED")
                cc.alignment = CENTER
            r += 1

    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].hidden = True
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 58
    ws.freeze_panes = f"{first_prov_col}{first_row}"
    protect(ws)
    return ws

def om_seed(slot, cp, qq):
    if slot > len(PROVS):
        return None
    pid = PROVS[slot - 1]["id"]
    avg = OM_SCORES.get(pid, {}).get(cp["id"])
    if avg is None:
        return None
    if qq["ty"] == "s":
        return avg
    return ("Settling in well, good rapport with staff." if avg >= 7
            else "Seems a bit withdrawn — not connecting with the team as hoped.")

def qp_seed(slot, cp, qq):
    if slot > len(PROVS):
        return None
    pid = PROVS[slot - 1]["id"]
    if qq.get("culture"):
        v = CII_SCORES.get(f"{pid}.{cp['id']}.{qq['qid']}")
        return int(v) if v is not None else None
    avg = SCORES.get(pid, {}).get(cp["id"])
    if avg is None:
        return None
    return avg if qq["ty"] == "s" else "Demo response"

# ═════════════════════════════════════════════════════════════════════════════
METRICS_HEADER_ROW = 12
METRICS_LAST_ROW = METRICS_HEADER_ROW + 40   # example + 39 blank entry rows

def build_metrics(wb, region):
    ws = wb.create_sheet(SHEETS["metrics"])
    home_link(ws)
    title_block(ws, region, "provider productivity metrics — IT integration landing zone")

    ws.merge_cells("A3:I4")
    note = ws["A3"]
    note.value = ("This sheet is the landing zone for automated provider metrics. IT: append one row per "
                  "provider per week to the table below — the column contract is documented on the Data "
                  "Dictionary sheet. Dashboards reference this table and light up automatically as rows land. "
                  "Until the feed exists, rows may be entered manually.")
    note.font = F(9, italic=True, color="1E40AF")
    note.fill = fill("EFF6FF")
    note.alignment = WRAP

    ws["A6"] = "Practice goals (from the onboarding program benchmarks)"
    ws["A6"].font = F(10, bold=True, color=NAVY)
    for i, h in enumerate(["Metric", "Unit", "Practice Goal", "Direction"], start=1):
        style_header_cell(ws.cell(row=7, column=i, value=h), region["accent"])
    for j, (name, unit, goal, better, _d) in enumerate(METRIC_COLS):
        r = 8 + j
        ws.cell(row=r, column=1, value=name).font = F(9)
        ws.cell(row=r, column=2, value=unit).font = F(9)
        ws.cell(row=r, column=3, value=goal).font = F(9)
        ws.cell(row=r, column=4, value=better).font = F(9)
        for cidx in range(1, 5):
            ws.cell(row=r, column=cidx).border = BOX

    headers = (["Week Ending", "Provider ID", "Provider Name"] +
               [m[0] for m in METRIC_COLS] + ["Source System"])
    for i, h in enumerate(headers, start=1):
        style_header_cell(ws.cell(row=METRICS_HEADER_ROW, column=i, value=h), region["accent"])
    widths = [12, 11, 24, 13, 13, 15, 13, 15, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Example row
    r = METRICS_HEADER_ROW + 1
    example_name = PROVS[0]["name"] if region["seed"] else "Dr. A. Example"
    example_id = PROVS[0]["id"] if region["seed"] else "EX-01"
    last_monday = TODAY - datetime.timedelta(days=TODAY.weekday() + 2)  # most recent Saturday-ish
    ws.cell(row=r, column=1, value=last_monday).number_format = "m/d/yyyy"
    ws.cell(row=r, column=2, value=example_id)
    ws.cell(row=r, column=3, value=example_name)
    for j, v in enumerate([16.5, 172, 26, 15.5, 7.0], start=4):
        ws.cell(row=r, column=j, value=v)
    ws.cell(row=r, column=9, value="EXAMPLE ROW — manually entered; replace with real data")
    ws.cell(row=r, column=9).font = F(8, italic=True, color="B45309")

    for rr in range(METRICS_HEADER_ROW + 1, METRICS_LAST_ROW + 1):
        for cc in range(1, 10):
            cell = ws.cell(row=rr, column=cc)
            cell.protection = UNLOCKED
            cell.border = BOX
            cell.font = cell.font.copy(name=FONT, size=9) if cell.font else F(9)
            if cc == 1:
                cell.number_format = "m/d/yyyy"

    ref = f"A{METRICS_HEADER_ROW}:I{METRICS_LAST_ROW}"
    table = Table(displayName="ProviderMetrics", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = f"A{METRICS_HEADER_ROW + 1}"
    protect(ws)
    return ws

# ═════════════════════════════════════════════════════════════════════════════
def build_dashboard(wb, region):
    ws = wb.create_sheet(SHEETS["dash"])
    home_link(ws)
    title_block(ws, region, "live program overview — every figure recomputes from the tracking sheets")

    roster = q(SHEETS["roster"])
    lastr = ROSTER_FIRST_ROW + N_SLOTS - 1
    name_rng = f"{roster}!$B${ROSTER_FIRST_ROW}:$B${lastr}"
    status_rng = f"{roster}!$I${ROSTER_FIRST_ROW}:$I${lastr}"
    qsheet = q(SHEETS["q"])

    tiles = [
        ("Providers", f"={nonblank_count(name_rng)}", "1F2937", "E5E7EB"),
        ("On Track", f'=COUNTIF({status_rng},"On Track")', "15803D", "DCFCE7"),
        ("Due", f'=COUNTIF({status_rng},"Due")', "B45309", "FEF3C7"),
        ("Overdue", f'=COUNTIF({status_rng},"Overdue")', "B91C1C", "FEE2E2"),
        ("Culture risk (C.I.I. < 6)", f'=COUNTIF({qsheet}!$E$4:$P$4,"<6")', "6D28D9", "EDE9FE"),
    ]
    for i, (label, formula, fcolor, bg) in enumerate(tiles):
        c1 = 1 + i * 3
        ws.merge_cells(start_row=4, start_column=c1, end_row=4, end_column=c1 + 1)
        ws.merge_cells(start_row=5, start_column=c1, end_row=6, end_column=c1 + 1)
        lab = ws.cell(row=4, column=c1, value=label)
        lab.font = Font(name=FONT, size=9, bold=True, color=fcolor)
        lab.fill = fill(bg)
        lab.alignment = CENTER
        val = ws.cell(row=5, column=c1, value=formula)
        val.font = Font(name=FONT, size=22, bold=True, color=fcolor)
        val.fill = fill(bg)
        val.alignment = CENTER

    hdrs = ["Provider", "Status", "Medical Director %", "Mentor Physician %",
            "Mentor APC %", "Office Manager Avg", "Questionnaire Avg",
            "Culture Index"] + [m[0] for m in METRIC_COLS]
    hr = 9
    for i, h in enumerate(hdrs, start=1):
        style_header_cell(ws.cell(row=hr, column=i, value=h), region["accent"])
    widths = [24, 10, 12, 12, 12, 12, 12, 10, 12, 12, 13, 12, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    md, phys, apc = q(SHEETS["md"]), q(SHEETS["phys"]), q(SHEETS["apc"])
    om_s = q(SHEETS["om"])
    met = q(SHEETS["metrics"])
    mrng = lambda col: f"{met}!${col}${METRICS_HEADER_ROW + 1}:${col}${METRICS_LAST_ROW}"

    for slot in range(1, N_SLOTS + 1):
        r = hr + slot
        rr = ROSTER_FIRST_ROW + slot - 1
        wide = prov_col(slot, "E")    # provider col on MD/OM/Q sheets
        narrow = prov_col(slot, "C")  # provider col on mentor sheets
        guard = f'IF({roster}!$B${rr}="","",'
        ws.cell(row=r, column=1, value=f'={roster}!$B${rr}').font = F(9, bold=True)
        ws.cell(row=r, column=2, value=f'={guard}{roster}!$I${rr})')
        ws.cell(row=r, column=3, value=f"={guard}{md}!{wide}$3)").number_format = "0%"
        ws.cell(row=r, column=4, value=f"={guard}{phys}!{narrow}$3)").number_format = "0%"
        ws.cell(row=r, column=5, value=f"={guard}{apc}!{narrow}$3)").number_format = "0%"
        ws.cell(row=r, column=6, value=f"={guard}{om_s}!{wide}$3)").number_format = "0.0"
        ws.cell(row=r, column=7, value=f"={guard}{qsheet}!{wide}$3)").number_format = "0.0"
        ws.cell(row=r, column=8, value=f"={guard}{qsheet}!{wide}$4)").number_format = "0.0"
        for j, col in enumerate("DEFGH", start=9):
            ws.cell(row=r, column=j, value=(
                f'={guard}IFERROR(AVERAGEIFS({mrng(col)},{mrng("B")},{roster}!$A${rr}),"—"))'
            )).number_format = "0.0"
        for cidx in range(1, 14):
            cell = ws.cell(row=r, column=cidx)
            cell.border = BOX
            if not cell.font or cell.font.name != FONT:
                cell.font = F(9)
            cell.alignment = CENTER if cidx > 1 else Alignment(vertical="center")

    body_last = hr + N_SLOTS
    ws.conditional_formatting.add(f"B{hr + 1}:B{body_last}", CellIsRule(
        operator="equal", formula=['"Overdue"'], fill=fill("FECACA"),
        font=Font(name=FONT, bold=True, color="B91C1C")))
    ws.conditional_formatting.add(f"B{hr + 1}:B{body_last}", CellIsRule(
        operator="equal", formula=['"Due"'], fill=fill("FEF3C7"),
        font=Font(name=FONT, bold=True, color="B45309")))
    ws.conditional_formatting.add(f"H{hr + 1}:H{body_last}", CellIsRule(
        operator="lessThan", formula=["6"], fill=fill("FECACA"),
        font=Font(name=FONT, bold=True, color="B91C1C")))

    ws.cell(row=body_last + 2, column=1,
            value=("Metrics columns show the average of all rows on the Provider Metrics sheet for each "
                   "provider; they read \"—\" until metric rows are entered or the IT feed lands.")
            ).font = F(9, italic=True, color="6B7280")
    protect(ws)
    return ws

# ═════════════════════════════════════════════════════════════════════════════
def build_data_dictionary(wb, region):
    ws = wb.create_sheet(SHEETS["dd"])
    home_link(ws)
    title_block(ws, region, "column contract for the IT metrics feed + workbook conventions")

    ws["A4"] = "Provider Metrics table — column contract (table name: ProviderMetrics)"
    ws["A4"].font = F(11, bold=True, color=NAVY)
    hdrs = ["Column", "Type", "Unit", "Expected Range", "Refresh", "Source (to be confirmed by IT)", "Notes"]
    for i, h in enumerate(hdrs, start=1):
        style_header_cell(ws.cell(row=5, column=i, value=h), region["accent"])
    rows = [
        ("Week Ending", "Date", "—", "Any Saturday", "Weekly", "Feed timestamp", "One row per provider per week."),
        ("Provider ID", "Text", "—", "Matches Roster column A", "Weekly", "Epic provider record (SER) or HR ID",
         "Join key — must match the Roster sheet exactly."),
        ("Provider Name", "Text", "—", "—", "Weekly", "Epic / HR", "Display only; the ID is the join key."),
    ]
    for name, unit, goal, better, desc in METRIC_COLS:
        rows.append((name, "Number", unit, "≥ 0", "Weekly",
                     "Epic Clarity / Caboodle extract — IT to map",
                     f"{desc} Practice goal: {goal} ({better.lower()})."))
    rows.append(("Source System", "Text", "—", "—", "Weekly", "Feed metadata",
                 "Name of the system that produced the row (or 'manual')."))
    for j, row in enumerate(rows):
        r = 6 + j
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = F(9)
            c.alignment = WRAP
            c.border = BOX

    r = 6 + len(rows) + 2
    ws.cell(row=r, column=1, value="Workbook conventions").font = F(11, bold=True, color=NAVY)
    conventions = [
        "Yellow cells are input cells — everything you are expected to type lives there. All other cells are formulas or reference content.",
        "✔ checkboxes: pick ✔ from the cell dropdown (or type it); delete to un-check.",
        "Scores are whole numbers 1–10. Anchors for 1 and 10 are printed with each question.",
        "Roster Status logic: a provider is Due when one phase behind the expected phase for their days in practice, Overdue when two or more behind.",
        "Sheets are protected without a password purely as a guardrail — Review > Unprotect Sheet lifts it.",
        "This workbook was generated from the meridian-os curriculum source. To change curriculum content, regenerate rather than editing item text by hand, or hand edits will be lost on the next regeneration.",
        PROGRAM_ADMINS + ".",
    ]
    if region["seed"]:
        conventions.insert(0, "THIS COPY CONTAINS A DEMO COHORT (9 fictional providers with seeded checks and scores) so every surface renders. Clear the Roster and tracking columns before live use, or keep it as the reference/demo copy.")
    else:
        conventions.insert(0, "The Roster contains one EXAMPLE row showing the expected format — replace it with your first real provider.")
    for j, text in enumerate(conventions):
        c = ws.cell(row=r + 1 + j, column=1, value="• " + text)
        c.font = F(9)
        c.alignment = WRAP
        ws.merge_cells(start_row=r + 1 + j, start_column=1, end_row=r + 1 + j, end_column=7)
        ws.row_dimensions[r + 1 + j].height = 26

    for col, w in zip("ABCDEFG", [26, 9, 9, 20, 9, 34, 52]):
        ws.column_dimensions[col].width = w
    protect(ws)
    return ws

# ═════════════════════════════════════════════════════════════════════════════
def build_home(wb, region):
    ws = wb.create_sheet(SHEETS["home"], 0)
    ws.sheet_view.showGridLines = False
    for col in range(1, 18):
        ws.column_dimensions[get_column_letter(col)].width = 9
    for row in range(1, 46):
        for col in range(1, 18):
            ws.cell(row=row, column=col).fill = fill(NAVY)

    ws.merge_cells("B3:P3")
    c = ws["B3"]
    c.value = "PROVIDER ONBOARDING & MENTORSHIP TRACKER"
    c.font = Font(name=FONT, size=11, bold=True, color="9CA3AF")
    ws.merge_cells("B4:P5")
    c = ws["B4"]
    c.value = region["name"]
    c.font = Font(name=FONT, size=30, bold=True, color="FFFFFF")
    c.alignment = Alignment(vertical="center")
    ws.merge_cells("B6:P6")
    c = ws["B6"]
    c.value = f"formerly {region['legacy']}  ·  Optum Tri-State"
    c.font = Font(name=FONT, size=11, italic=True, color="9CA3AF")

    ws.merge_cells("B8:P8")
    c = ws["B8"]
    c.value = f"   {region['director']}  —  Primary Care Medical Director"
    c.font = Font(name=FONT, size=12, bold=True, color="FFFFFF")
    c.fill = fill(region["accent"])
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[8].height = 26

    roster = q(SHEETS["roster"])
    lastr = ROSTER_FIRST_ROW + N_SLOTS - 1
    md, phys, apc = q(SHEETS["md"]), q(SHEETS["phys"]), q(SHEETS["apc"])
    om_s, qs, met = q(SHEETS["om"]), q(SHEETS["q"]), q(SHEETS["metrics"])
    avg_txt = lambda sheet, rng: f'=IFERROR(TEXT(AVERAGE({sheet}!{rng}),"0%")&" average completion","Not started")'
    score_txt = lambda sheet, row: f'=IFERROR(TEXT(AVERAGE({sheet}!E{row}:P{row}),"0.0")&" average score","Not started")'

    tiles = [
        ("Roster", "roster",
         f'={nonblank_count(f"{roster}!$B${ROSTER_FIRST_ROW}:$B${lastr}")}&IF('
         f'{nonblank_count(f"{roster}!$B${ROSTER_FIRST_ROW}:$B${lastr}")}=1,'
         f'" provider onboarding"," providers onboarding")'),
        ("Medical Director Curriculum", "md", avg_txt(md, "$E$3:$P$3")),
        ("Mentor — Physician Track", "phys", avg_txt(phys, "$C$3:$N$3")),
        ("Mentor — APC Track", "apc", avg_txt(apc, "$C$3:$N$3")),
        ("Office Manager Reviews", "om", score_txt(om_s, 3)),
        ("Provider Questionnaires", "q", score_txt(qs, 3)),
        ("Provider Metrics", "metrics",
         f'={nonblank_count(f"{met}!$B${METRICS_HEADER_ROW + 1}:$B${METRICS_LAST_ROW}")}&IF('
         f'{nonblank_count(f"{met}!$B${METRICS_HEADER_ROW + 1}:$B${METRICS_LAST_ROW}")}=1,'
         f'" metric row landed"," metric rows landed")'),
        ("Dashboard", "dash",
         f'=COUNTIF({roster}!$I${ROSTER_FIRST_ROW}:$I${lastr},"Overdue")&IF(COUNTIF('
         f'{roster}!$I${ROSTER_FIRST_ROW}:$I${lastr},"Overdue")=1," provider overdue"," providers overdue")'),
    ]
    starts = ["B", "F", "J", "N"]
    for i, (label, key, stat) in enumerate(tiles):
        color = MONDRIAN[i % 4]
        dark_text = color == "EDE8DF"
        col0 = ord(starts[i % 4]) - ord("A") + 1
        row0 = 11 if i < 4 else 18
        ws.merge_cells(start_row=row0, start_column=col0, end_row=row0 + 2, end_column=col0 + 2)
        ws.merge_cells(start_row=row0 + 3, start_column=col0, end_row=row0 + 4, end_column=col0 + 2)
        name_cell = ws.cell(row=row0, column=col0)
        name_cell.value = label
        name_cell.font = Font(name=FONT, size=13, bold=True,
                              color=NAVY if dark_text else "FFFFFF")
        name_cell.alignment = CENTER_WRAP
        name_cell.hyperlink = Hyperlink(ref=name_cell.coordinate,
                                        location=f"{q(SHEETS[key])}!A1",
                                        tooltip=f"Open {SHEETS[key]}")
        stat_cell = ws.cell(row=row0 + 3, column=col0)
        stat_cell.value = stat
        stat_cell.font = Font(name=FONT, size=9, italic=True,
                              color="374151" if dark_text else "E5E7EB")
        stat_cell.alignment = CENTER_WRAP
        for rr in range(row0, row0 + 5):
            for cc2 in range(col0, col0 + 3):
                ws.cell(row=rr, column=cc2).fill = fill(color)
        ws.row_dimensions[row0].height = 20

    ws.merge_cells("B25:P26")
    c = ws["B25"]
    c.value = ("Click a tile to open its sheet — every sheet has a ⌂ Home link in the top-left corner. "
               "Yellow cells are where you type; everything else computes itself. "
               "Conventions, the reset checklist, and the IT metrics contract live on the Data Dictionary sheet.")
    c.font = Font(name=FONT, size=10, italic=True, color="9CA3AF")
    c.alignment = WRAP
    ws.merge_cells("B28:P28")
    c = ws["B28"]
    c.value = PROGRAM_ADMINS
    c.font = Font(name=FONT, size=9, color="6B7280")
    ws.merge_cells("B29:P29")
    c = ws["B29"]
    c.value = "For authorized Optum staff only · No patient information belongs in this workbook"
    c.font = Font(name=FONT, size=9, color="6B7280")
    return ws

# ═════════════════════════════════════════════════════════════════════════════
def build_region(region):
    wb = Workbook()
    wb.remove(wb.active)
    build_lists(wb)
    build_roster(wb, region)
    build_md_curriculum(wb, region)
    build_mentor_sheet(wb, region, "phys", PHYS,
                       "weekly/monthly mentor check-ins — all providers, Week 0 through Month 12",
                       lambda p: True)
    build_mentor_sheet(wb, region, "apc", APC,
                       "quarterly mentor check-ins — Advanced Practice Clinicians (NP/PA) only, Months 15–24",
                       lambda p: p["role"] in ("NP", "PA"),
                       credential_gate=("NP", "PA"))
    build_question_sheet(wb, region, "om", OM,
                         "office manager review checkpoints — scores 1–10 with a verbatim provider check-in",
                         om_seed, culture_aware=False)
    build_question_sheet(wb, region, "q", QP,
                         "provider self-report questionnaires — includes the Culture Integration Index items",
                         qp_seed, culture_aware=True)
    build_metrics(wb, region)
    build_dashboard(wb, region)
    build_data_dictionary(wb, region)
    build_home(wb, region)

    # Tab order/colors: Home first (already index 0 via create position).
    order = [SHEETS[k] for k in ("home", "roster", "dash", "md", "phys", "apc",
                                 "om", "q", "metrics", "dd", "lists")]
    wb._sheets.sort(key=lambda s: order.index(s.title))
    for s in wb._sheets:
        s.sheet_properties.tabColor = region["accent"]
    wb.active = 0

    # openpyxl writes formulas with no cached values, so every computed cell
    # would read as blank until something recalculates. Force Excel to do a
    # full recalculation the first time the workbook is opened.
    wb.calculation.fullCalcOnLoad = True

    out = OUT_DIR / f"Onboarding-Tracker-{region['name'].replace(' ', '-')}.xlsx"
    wb.save(out)
    print(f"wrote {out}")
    return out

if __name__ == "__main__":
    for region in REGIONS:
        build_region(region)
